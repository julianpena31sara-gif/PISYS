from flask import Flask, render_template, request, redirect, url_for, flash, make_response, send_file
from flask_sqlalchemy import SQLAlchemy
from flask_login import LoginManager, UserMixin, login_user, logout_user, login_required, current_user
from functools import wraps
from werkzeug.security import generate_password_hash, check_password_hash
from datetime import datetime
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.lib.units import cm
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, HRFlowable
from reportlab.lib.enums import TA_CENTER, TA_RIGHT
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

app = Flask(__name__)
app.config['SECRET_KEY'] = 'clave-super-secreta-cambiala-en-produccion'
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///inventario.db'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

db = SQLAlchemy(app)

# Flask-Login: maneja sesiones de usuario automáticamente
login_manager = LoginManager(app)
login_manager.login_view        = 'login'           # ruta a la que redirige si no hay sesión
login_manager.login_message     = 'Inicia sesión para continuar.'
login_manager.login_message_category = 'warning'


# ---------------------------------------------------------------------------
# MODELOS
# ---------------------------------------------------------------------------

class Producto(db.Model):
    __tablename__ = 'producto'

    id             = db.Column(db.Integer, primary_key=True)
    nombre         = db.Column(db.String(100), nullable=False)
    descripcion    = db.Column(db.Text, nullable=True)
    precio         = db.Column(db.Float, nullable=False, default=0.0)
    cantidad       = db.Column(db.Integer, nullable=False, default=0)
    categoria      = db.Column(db.String(50), nullable=True, default='General')
    fecha_creacion = db.Column(db.DateTime, default=datetime.utcnow)

    def __repr__(self):
        return f'<Producto {self.nombre}>'

    def estado_stock(self):
        if self.cantidad < 5:  return 'critico'
        if self.cantidad < 10: return 'bajo'
        return 'normal'

    def valor_total(self):
        return self.precio * self.cantidad


class Movimiento(db.Model):
    __tablename__ = 'movimiento'

    id                = db.Column(db.Integer, primary_key=True)
    producto_id       = db.Column(db.Integer, db.ForeignKey('producto.id'), nullable=False)
    tipo              = db.Column(db.String(20), nullable=False, default='ajuste')  # entrada | salida | ajuste | creacion
    cantidad_cambio   = db.Column(db.Integer, nullable=False)
    cantidad_anterior = db.Column(db.Integer, nullable=False)
    cantidad_nueva    = db.Column(db.Integer, nullable=False)
    nota              = db.Column(db.String(200), nullable=True)
    fecha             = db.Column(db.DateTime, default=datetime.utcnow)

    producto = db.relationship('Producto', backref='movimientos')

    def __repr__(self):
        signo = '+' if self.cantidad_cambio >= 0 else ''
        return f'<Movimiento {self.producto_id} {signo}{self.cantidad_cambio}>'


class Usuario(UserMixin, db.Model):
    """
    Modelo de usuario con roles RBAC.
    Hereda de UserMixin para que Flask-Login funcione automáticamente
    (provee is_authenticated, is_active, get_id(), etc.)
    Roles válidos: 'admin' | 'operador' | 'auditor'
    """
    __tablename__ = 'usuario'

    id             = db.Column(db.Integer, primary_key=True)
    nombre         = db.Column(db.String(80), nullable=False)
    email          = db.Column(db.String(120), unique=True, nullable=False)
    password_hash  = db.Column(db.String(256), nullable=False)
    rol            = db.Column(db.String(20), nullable=False, default='operador')
    activo         = db.Column(db.Boolean, default=True)
    fecha_creacion = db.Column(db.DateTime, default=datetime.utcnow)

    def set_password(self, password):
        """Hashea y guarda la contraseña. Nunca guardamos texto plano."""
        self.password_hash = generate_password_hash(password)

    def check_password(self, password):
        """Verifica la contraseña contra el hash guardado."""
        return check_password_hash(self.password_hash, password)

    def tiene_rol(self, *roles):
        """Devuelve True si el usuario tiene alguno de los roles indicados."""
        return self.rol in roles

    def __repr__(self):
        return f'<Usuario {self.email} [{self.rol}]>'


@login_manager.user_loader
def cargar_usuario(user_id):
    """Flask-Login llama esto en cada request para reconstruir current_user."""
    return Usuario.query.get(int(user_id))


# ---------------------------------------------------------------------------
# DECORADORES DE ROLES
# ---------------------------------------------------------------------------

def requiere_rol(*roles):
    """
    Decorador que restringe una ruta a uno o más roles específicos.
    Uso: @requiere_rol('admin') o @requiere_rol('admin', 'operador')
    Siempre va DESPUÉS de @login_required.
    """
    def decorador(f):
        @wraps(f)
        def wrapper(*args, **kwargs):
            if not current_user.tiene_rol(*roles):
                flash('⛔ No tienes permisos para acceder a esa sección.', 'danger')
                return redirect(url_for('inicio'))
            return f(*args, **kwargs)
        return wrapper
    return decorador


# ---------------------------------------------------------------------------
# HELPERS
# ---------------------------------------------------------------------------

def registrar_movimiento(producto, cantidad_anterior, tipo='ajuste', nota=None):
    """Registra un movimiento de stock. El commit lo hace el llamador."""
    cambio = producto.cantidad - cantidad_anterior
    if cambio == 0 and tipo != 'creacion':
        return
    db.session.add(Movimiento(
        producto_id       = producto.id,
        tipo              = tipo,
        cantidad_cambio   = cambio,
        cantidad_anterior = cantidad_anterior,
        cantidad_nueva    = producto.cantidad,
        nota              = nota
    ))


def _estilos_pdf():
    """Estilos ReportLab reutilizables en todos los reportes."""
    base = getSampleStyleSheet()
    return {
        'titulo':    ParagraphStyle('titulo',    parent=base['Title'],
                                    fontSize=20, textColor=colors.HexColor('#1E3A5F'), spaceAfter=4),
        'subtitulo': ParagraphStyle('subtitulo', parent=base['Normal'],
                                    fontSize=10, textColor=colors.HexColor('#64748B'), spaceAfter=2),
        'seccion':   ParagraphStyle('seccion',   parent=base['Normal'],
                                    fontSize=12, fontName='Helvetica-Bold',
                                    textColor=colors.HexColor('#1E3A5F'), spaceBefore=14, spaceAfter=6),
        'normal':    ParagraphStyle('normal',    parent=base['Normal'],
                                    fontSize=9,  textColor=colors.HexColor('#374151')),
        'pie':       ParagraphStyle('pie',       parent=base['Normal'],
                                    fontSize=8,  textColor=colors.HexColor('#9CA3AF'), alignment=TA_CENTER),
        'derecha':   ParagraphStyle('derecha',   parent=base['Normal'],
                                    fontSize=9,  textColor=colors.HexColor('#374151'), alignment=TA_RIGHT),
    }


def _encabezado_pdf(elementos, estilos, titulo, subtitulo):
    """Encabezado estándar de todos los reportes."""
    elementos.append(Paragraph('PISYS', estilos['titulo']))
    elementos.append(Paragraph(titulo, estilos['seccion']))
    elementos.append(Paragraph(
        f'{subtitulo} &nbsp;·&nbsp; Generado: {datetime.now().strftime("%d/%m/%Y %H:%M")}',
        estilos['subtitulo']
    ))
    elementos.append(HRFlowable(width='100%', thickness=1,
                                color=colors.HexColor('#2563EB'), spaceAfter=12))


def _estilo_tabla_base():
    """Encabezado azul oscuro + filas alternas gris claro."""
    return TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0),  colors.HexColor('#1E3A5F')),
        ('TEXTCOLOR',  (0, 0), (-1, 0),  colors.white),
        ('FONTNAME',   (0, 0), (-1, 0),  'Helvetica-Bold'),
        ('FONTSIZE',   (0, 0), (-1, 0),  9),
        ('FONTNAME',   (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE',   (0, 1), (-1, -1), 8),
        ('ROWBACKGROUND', (0, 2), (-1, -1), colors.HexColor('#F8FAFC')),
        ('GRID',       (0, 0), (-1, -1), 0.4, colors.HexColor('#E2E8F0')),
        ('PADDING',    (0, 0), (-1, -1), 6),
        ('VALIGN',     (0, 0), (-1, -1), 'MIDDLE'),
    ])


def _enviar_pdf(buffer, filename):
    """Convierte el buffer en una respuesta HTTP con tipo PDF."""
    buffer.seek(0)
    response = make_response(buffer.read())
    response.headers['Content-Type']        = 'application/pdf'
    response.headers['Content-Disposition'] = f'inline; filename={filename}'
    return response


CATEGORIAS = ['Electrónica', 'Papelería', 'Mobiliario', 'Limpieza',
              'Alimentos', 'Ropa', 'Herramientas', 'General']


# ---------------------------------------------------------------------------
# RUTAS — INICIO
# ---------------------------------------------------------------------------

@app.route('/')
@login_required
def inicio():
    productos = Producto.query.all()
    return render_template('inicio.html',
        total_productos        = Producto.query.count(),
        alertas_criticas       = [p for p in productos if p.estado_stock() == 'critico'],
        alertas_bajas          = [p for p in productos if p.estado_stock() == 'bajo'],
        valor_total_inventario = sum(p.valor_total() for p in productos)
    )


# ---------------------------------------------------------------------------
# RUTAS — PRODUCTOS (CRUD)
# ---------------------------------------------------------------------------

@app.route('/productos')
@login_required
def lista_productos():
    buscar    = request.args.get('buscar', '').strip()
    categoria = request.args.get('categoria', '')
    pagina    = request.args.get('pagina', 1, type=int)

    consulta = Producto.query
    if buscar:
        consulta = consulta.filter(Producto.nombre.ilike(f'%{buscar}%'))
    if categoria:
        consulta = consulta.filter(Producto.categoria == categoria)

    productos  = consulta.order_by(Producto.fecha_creacion.desc()) \
                         .paginate(page=pagina, per_page=10, error_out=False)
    categorias = [c[0] for c in db.session.query(Producto.categoria)
                                          .distinct().order_by(Producto.categoria).all() if c[0]]

    return render_template('productos.html',
                           productos=productos, categorias=categorias,
                           buscar=buscar, categoria_activa=categoria)


@app.route('/productos/<int:id>')
@login_required
def detalle_producto(id):
    return render_template('detalle.html', producto=Producto.query.get_or_404(id))


@app.route('/productos/nuevo', methods=['GET', 'POST'])
@requiere_rol('admin', 'operador')
@login_required
def nuevo_producto():
    if request.method == 'POST':
        nombre      = request.form.get('nombre', '').strip()
        descripcion = request.form.get('descripcion', '').strip()
        categoria   = request.form.get('categoria', 'General').strip()

        try:
            precio   = float(request.form.get('precio', 0))
            cantidad = int(request.form.get('cantidad', 0))
        except ValueError:
            flash('❌ El precio y la cantidad deben ser números válidos.', 'danger')
            return render_template('formulario.html', titulo='Nuevo Producto',
                                   categorias=CATEGORIAS, producto=None)

        errores = []
        if not nombre:           errores.append('El nombre es obligatorio.')
        if len(nombre) > 100:    errores.append('El nombre no puede superar 100 caracteres.')
        if precio < 0:           errores.append('El precio no puede ser negativo.')
        if cantidad < 0:         errores.append('La cantidad no puede ser negativa.')

        if errores:
            for e in errores: flash(f'❌ {e}', 'danger')
            return render_template('formulario.html', titulo='Nuevo Producto',
                                   categorias=CATEGORIAS,
                                   producto={'nombre': nombre, 'descripcion': descripcion,
                                             'precio': precio, 'cantidad': cantidad, 'categoria': categoria})

        nuevo = Producto(nombre=nombre, descripcion=descripcion,
                         precio=precio, cantidad=cantidad, categoria=categoria)
        db.session.add(nuevo)
        db.session.flush()  # obtiene el ID antes del commit

        if nuevo.cantidad > 0:
            registrar_movimiento(nuevo, 0, tipo='creacion', nota='Stock inicial')

        db.session.commit()
        flash(f'✅ Producto "{nombre}" creado exitosamente.', 'success')
        return redirect(url_for('lista_productos'))

    return render_template('formulario.html', titulo='Nuevo Producto',
                           categorias=CATEGORIAS, producto=None)


@app.route('/productos/editar/<int:id>', methods=['GET', 'POST'])
@requiere_rol('admin', 'operador')
@login_required
def editar_producto(id):
    producto = Producto.query.get_or_404(id)

    if request.method == 'POST':
        nombre      = request.form.get('nombre', '').strip()
        descripcion = request.form.get('descripcion', '').strip()
        categoria   = request.form.get('categoria', 'General').strip()

        try:
            precio   = float(request.form.get('precio', 0))
            cantidad = int(request.form.get('cantidad', 0))
        except ValueError:
            flash('❌ El precio y la cantidad deben ser números válidos.', 'danger')
            return render_template('formulario.html', titulo=f'Editar: {producto.nombre}',
                                   categorias=CATEGORIAS, producto=producto)

        errores = []
        if not nombre:        errores.append('El nombre es obligatorio.')
        if len(nombre) > 100: errores.append('El nombre no puede superar 100 caracteres.')
        if precio < 0:        errores.append('El precio no puede ser negativo.')
        if cantidad < 0:      errores.append('La cantidad no puede ser negativa.')

        if errores:
            for e in errores: flash(f'❌ {e}', 'danger')
            return render_template('formulario.html', titulo=f'Editar: {producto.nombre}',
                                   categorias=CATEGORIAS, producto=producto)

        cantidad_anterior    = producto.cantidad
        producto.nombre      = nombre
        producto.descripcion = descripcion
        producto.precio      = precio
        producto.cantidad    = cantidad
        producto.categoria   = categoria

        if cantidad != cantidad_anterior:
            registrar_movimiento(producto, cantidad_anterior,
                                 tipo='ajuste', nota='Modificación desde formulario')

        db.session.commit()
        flash(f'✅ Producto "{nombre}" actualizado correctamente.', 'success')
        return redirect(url_for('detalle_producto', id=producto.id))

    return render_template('formulario.html', titulo=f'Editar: {producto.nombre}',
                           categorias=CATEGORIAS, producto=producto)


@app.route('/productos/eliminar/<int:id>', methods=['POST'])
@requiere_rol('admin')
@login_required
def eliminar_producto(id):
    producto = Producto.query.get_or_404(id)
    nombre   = producto.nombre
    Movimiento.query.filter_by(producto_id=id).delete()
    db.session.delete(producto)
    db.session.commit()
    flash(f'🗑️ Producto "{nombre}" eliminado.', 'warning')
    return redirect(url_for('lista_productos'))


@app.route('/productos/ajustar-stock/<int:id>', methods=['POST'])
@requiere_rol('admin', 'operador')
@login_required
def ajustar_stock(id):
    producto = Producto.query.get_or_404(id)

    try:
        cantidad_ajuste = int(request.form.get('cantidad_ajuste', 0))
    except ValueError:
        flash('❌ La cantidad debe ser un número entero.', 'danger')
        return redirect(url_for('detalle_producto', id=id))

    if cantidad_ajuste <= 0:
        flash('❌ La cantidad debe ser mayor a 0.', 'danger')
        return redirect(url_for('detalle_producto', id=id))

    cantidad_anterior = producto.cantidad
    operacion         = request.form.get('operacion', 'sumar')

    if operacion == 'sumar':
        producto.cantidad += cantidad_ajuste
        tipo, nota = 'entrada', f'Entrada de {cantidad_ajuste} unidades'
    else:
        if producto.cantidad - cantidad_ajuste < 0:
            flash(f'❌ No puedes restar {cantidad_ajuste}. Solo hay {producto.cantidad} en stock.', 'danger')
            return redirect(url_for('detalle_producto', id=id))
        producto.cantidad -= cantidad_ajuste
        tipo, nota = 'salida', f'Salida de {cantidad_ajuste} unidades'

    registrar_movimiento(producto, cantidad_anterior, tipo=tipo, nota=nota)
    db.session.commit()
    flash(f'✅ Stock: {cantidad_anterior} → {producto.cantidad} unidades.', 'success')
    return redirect(url_for('detalle_producto', id=id))


# ---------------------------------------------------------------------------
# RUTAS — HISTORIAL
# ---------------------------------------------------------------------------

@app.route('/historial')
@requiere_rol('admin', 'auditor')
@login_required
def historial():
    pagina      = request.args.get('pagina', 1, type=int)
    tipo_filtro = request.args.get('tipo', '')
    fecha_desde = request.args.get('fecha_desde', '')
    fecha_hasta = request.args.get('fecha_hasta', '')

    consulta = Movimiento.query

    if tipo_filtro:
        consulta = consulta.filter(Movimiento.tipo == tipo_filtro)

    # Convertir strings a datetime y aplicar filtros de fecha
    dt_desde = dt_hasta = None
    if fecha_desde:
        try:
            dt_desde = datetime.strptime(fecha_desde, '%Y-%m-%d')
            consulta = consulta.filter(Movimiento.fecha >= dt_desde)
        except ValueError:
            flash('❌ Fecha "desde" inválida.', 'danger')

    if fecha_hasta:
        try:
            # Incluir todo el día hasta (hasta las 23:59:59)
            dt_hasta = datetime.strptime(fecha_hasta, '%Y-%m-%d').replace(
                hour=23, minute=59, second=59)
            consulta = consulta.filter(Movimiento.fecha <= dt_hasta)
        except ValueError:
            flash('❌ Fecha "hasta" inválida.', 'danger')

    movimientos = consulta.order_by(Movimiento.fecha.desc()) \
                          .paginate(page=pagina, per_page=20, error_out=False)

    return render_template('historial.html',
                           movimientos=movimientos,
                           tipo_filtro=tipo_filtro,
                           fecha_desde=fecha_desde,
                           fecha_hasta=fecha_hasta)


# ---------------------------------------------------------------------------
# RUTAS — REPORTES PDF (Semanas 7, 8 y 9)
# ---------------------------------------------------------------------------

@app.route('/reportes')
@requiere_rol('admin', 'auditor')
@login_required
def reportes():
    productos = Producto.query.all()
    categorias = [c[0] for c in db.session.query(Producto.categoria)
                                          .distinct().order_by(Producto.categoria).all() if c[0]]
    return render_template('reportes.html',
        total_productos = len(productos),
        valor_total     = sum(p.valor_total() for p in productos),
        criticos        = len([p for p in productos if p.estado_stock() == 'critico']),
        bajos           = len([p for p in productos if p.estado_stock() == 'bajo']),
        categorias      = categorias
    )


@app.route('/reportes/inventario')
@requiere_rol('admin', 'auditor')
@login_required
def reporte_inventario():
    """Semana 8 — PDF completo de inventario con colores por estado y resumen."""
    categoria_filtro = request.args.get('categoria', '')

    consulta = Producto.query
    if categoria_filtro:
        consulta = consulta.filter(Producto.categoria == categoria_filtro)
    productos = consulta.order_by(Producto.categoria, Producto.nombre).all()

    buffer  = io.BytesIO()
    doc     = SimpleDocTemplate(buffer, pagesize=letter,
                                leftMargin=2*cm, rightMargin=2*cm,
                                topMargin=2*cm,  bottomMargin=2*cm)
    estilos   = _estilos_pdf()
    elementos = []

    subtitulo = f'Categoría: {categoria_filtro}' if categoria_filtro else 'Todos los productos'
    _encabezado_pdf(elementos, estilos, 'Reporte de Inventario', subtitulo)

    # Tabla principal
    filas = [['#', 'Producto', 'Categoría', 'Precio', 'Stock', 'Valor total', 'Estado']]
    for p in productos:
        estado = {'critico': 'CRÍTICO', 'bajo': 'BAJO', 'normal': 'Normal'}[p.estado_stock()]
        filas.append([str(p.id), Paragraph(p.nombre, estilos['normal']),
                      p.categoria or '—', f'${p.precio:,.0f}',
                      str(p.cantidad), f'${p.valor_total():,.0f}', estado])

    tabla = Table(filas,
                  colWidths=[1*cm, 5.5*cm, 3*cm, 2.5*cm, 1.5*cm, 2.5*cm, 2*cm],
                  repeatRows=1)
    estilo = _estilo_tabla_base()

    # Colorear filas según estado de stock
    for i, p in enumerate(productos, start=1):
        if p.estado_stock() == 'critico':
            estilo.add('BACKGROUND', (0, i), (-1, i), colors.HexColor('#FEF2F2'))
            estilo.add('TEXTCOLOR',  (6, i), (6, i),  colors.HexColor('#DC2626'))
            estilo.add('FONTNAME',   (6, i), (6, i),  'Helvetica-Bold')
        elif p.estado_stock() == 'bajo':
            estilo.add('BACKGROUND', (0, i), (-1, i), colors.HexColor('#FFFBEB'))
            estilo.add('TEXTCOLOR',  (6, i), (6, i),  colors.HexColor('#D97706'))
            estilo.add('FONTNAME',   (6, i), (6, i),  'Helvetica-Bold')
    tabla.setStyle(estilo)
    elementos.append(tabla)

    # Resumen ejecutivo
    elementos.append(Spacer(1, 0.5*cm))
    elementos.append(HRFlowable(width='100%', thickness=0.5,
                                color=colors.HexColor('#E2E8F0'), spaceAfter=8))
    criticos = len([p for p in productos if p.estado_stock() == 'critico'])
    bajos    = len([p for p in productos if p.estado_stock() == 'bajo'])
    resumen  = Table(
        [['Total productos', 'Valor inventario', 'Críticos', 'Stock bajo'],
         [str(len(productos)), f'${sum(p.valor_total() for p in productos):,.0f}',
          str(criticos), str(bajos)]],
        colWidths=[4.5*cm] * 4
    )
    resumen.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#F1F5F9')),
        ('FONTNAME',   (0, 0), (-1, -1), 'Helvetica-Bold'),
        ('FONTSIZE',   (0, 0), (-1, -1), 9),
        ('TEXTCOLOR',  (0, 1), (-1, 1),  colors.HexColor('#1E3A5F')),
        ('ALIGN',      (0, 0), (-1, -1), 'CENTER'),
        ('GRID',       (0, 0), (-1, -1), 0.4, colors.HexColor('#E2E8F0')),
        ('PADDING',    (0, 0), (-1, -1), 8),
    ]))
    elementos.append(resumen)
    elementos.append(Spacer(1, 0.3*cm))
    elementos.append(Paragraph('PISYS — Reporte generado automáticamente', estilos['pie']))

    doc.build(elementos)
    return _enviar_pdf(buffer, 'inventario.pdf')


@app.route('/reportes/alertas')
@requiere_rol('admin', 'auditor')
@login_required
def reporte_alertas():
    """Semana 8 — PDF de productos con stock crítico o bajo."""
    productos_alerta = Producto.query.filter(Producto.cantidad < 10) \
                                     .order_by(Producto.cantidad).all()

    buffer  = io.BytesIO()
    doc     = SimpleDocTemplate(buffer, pagesize=letter,
                                leftMargin=2*cm, rightMargin=2*cm,
                                topMargin=2*cm,  bottomMargin=2*cm)
    estilos   = _estilos_pdf()
    elementos = []

    _encabezado_pdf(elementos, estilos, 'Reporte de Alertas de Stock',
                    f'{len(productos_alerta)} producto(s) requieren atención')

    if not productos_alerta:
        elementos.append(Paragraph('✅ No hay productos con stock bajo o crítico.', estilos['normal']))
    else:
        filas = [['#', 'Producto', 'Categoría', 'Stock', 'Estado', 'Valor en riesgo']]
        for p in productos_alerta:
            estado = 'CRÍTICO' if p.estado_stock() == 'critico' else 'BAJO'
            filas.append([str(p.id), Paragraph(p.nombre, estilos['normal']),
                          p.categoria or '—', str(p.cantidad), estado,
                          f'${p.valor_total():,.0f}'])

        tabla = Table(filas,
                      colWidths=[1*cm, 6*cm, 3*cm, 2*cm, 2.5*cm, 3*cm],
                      repeatRows=1)
        estilo = _estilo_tabla_base()
        for i, p in enumerate(productos_alerta, start=1):
            c_bg   = colors.HexColor('#FEF2F2') if p.estado_stock() == 'critico' else colors.HexColor('#FFFBEB')
            c_text = colors.HexColor('#DC2626') if p.estado_stock() == 'critico' else colors.HexColor('#D97706')
            estilo.add('BACKGROUND', (0, i), (-1, i), c_bg)
            estilo.add('TEXTCOLOR',  (4, i), (4, i),  c_text)
            estilo.add('FONTNAME',   (4, i), (4, i),  'Helvetica-Bold')
        tabla.setStyle(estilo)
        elementos.append(tabla)

    elementos.append(Spacer(1, 0.5*cm))
    elementos.append(Paragraph('PISYS — Reporte de alertas', estilos['pie']))
    doc.build(elementos)
    return _enviar_pdf(buffer, 'alertas.pdf')


@app.route('/reportes/historial')
@requiere_rol('admin', 'auditor')
@login_required
def reporte_historial():
    """PDF del historial de movimientos con filtros de tipo y fechas."""
    tipo_filtro = request.args.get('tipo', '')
    fecha_desde = request.args.get('fecha_desde', '')
    fecha_hasta = request.args.get('fecha_hasta', '')

    consulta = Movimiento.query

    if tipo_filtro:
        consulta = consulta.filter(Movimiento.tipo == tipo_filtro)

    dt_desde = dt_hasta = None
    if fecha_desde:
        try:
            dt_desde = datetime.strptime(fecha_desde, '%Y-%m-%d')
            consulta = consulta.filter(Movimiento.fecha >= dt_desde)
        except ValueError:
            pass

    if fecha_hasta:
        try:
            dt_hasta = datetime.strptime(fecha_hasta, '%Y-%m-%d').replace(
                hour=23, minute=59, second=59)
            consulta = consulta.filter(Movimiento.fecha <= dt_hasta)
        except ValueError:
            pass

    movimientos = consulta.order_by(Movimiento.fecha.desc()).all()

    # Construir subtítulo descriptivo con los filtros activos
    partes = []
    if tipo_filtro:
        partes.append(f'Tipo: {tipo_filtro.capitalize()}')
    if fecha_desde:
        partes.append(f'Desde: {dt_desde.strftime("%d/%m/%Y")}')
    if fecha_hasta:
        partes.append(f'Hasta: {dt_hasta.strftime("%d/%m/%Y")}')
    subtitulo = ' · '.join(partes) if partes else 'Todos los movimientos'

    buffer  = io.BytesIO()
    doc     = SimpleDocTemplate(buffer, pagesize=letter,
                                leftMargin=2*cm, rightMargin=2*cm,
                                topMargin=2*cm,  bottomMargin=2*cm)
    estilos   = _estilos_pdf()
    elementos = []

    _encabezado_pdf(elementos, estilos, 'Historial de Movimientos', subtitulo)

    if not movimientos:
        elementos.append(Paragraph('No hay movimientos para el rango seleccionado.', estilos['normal']))
    else:
        filas = [['Fecha', 'Producto', 'Tipo', 'Antes', 'Cambio', 'Después', 'Nota']]
        for m in movimientos:
            signo  = '+' if m.cantidad_cambio > 0 else ''
            cambio = f'{signo}{m.cantidad_cambio}'
            filas.append([
                m.fecha.strftime('%d/%m/%Y %H:%M'),
                Paragraph(m.producto.nombre, estilos['normal']),
                m.tipo.capitalize(),
                str(m.cantidad_anterior),
                cambio,
                str(m.cantidad_nueva),
                Paragraph(m.nota or '—', estilos['normal']),
            ])

        tabla = Table(filas,
                      colWidths=[3*cm, 4.5*cm, 2*cm, 1.5*cm, 1.8*cm, 1.8*cm, 3.4*cm],
                      repeatRows=1)
        estilo = _estilo_tabla_base()
        for i, m in enumerate(movimientos, start=1):
            if m.cantidad_cambio > 0:
                estilo.add('TEXTCOLOR', (4, i), (4, i), colors.HexColor('#16A34A'))
                estilo.add('FONTNAME',  (4, i), (4, i), 'Helvetica-Bold')
            elif m.cantidad_cambio < 0:
                estilo.add('TEXTCOLOR', (4, i), (4, i), colors.HexColor('#DC2626'))
                estilo.add('FONTNAME',  (4, i), (4, i), 'Helvetica-Bold')
        tabla.setStyle(estilo)
        elementos.append(tabla)

        elementos.append(Spacer(1, 0.5*cm))
        elementos.append(Paragraph('Resumen', estilos['seccion']))
        entradas = sum(m.cantidad_cambio for m in movimientos if m.cantidad_cambio > 0)
        salidas  = sum(abs(m.cantidad_cambio) for m in movimientos if m.cantidad_cambio < 0)
        resumen  = Table(
            [['Total movimientos', 'Unidades ingresadas', 'Unidades salidas', 'Balance neto'],
             [str(len(movimientos)), f'+{entradas}', f'-{salidas}', f'{entradas - salidas:+}']],
            colWidths=[4.5*cm] * 4
        )
        resumen.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#F1F5F9')),
            ('FONTNAME',   (0, 0), (-1, -1), 'Helvetica-Bold'),
            ('FONTSIZE',   (0, 0), (-1, -1), 9),
            ('ALIGN',      (0, 0), (-1, -1), 'CENTER'),
            ('GRID',       (0, 0), (-1, -1), 0.4, colors.HexColor('#E2E8F0')),
            ('PADDING',    (0, 0), (-1, -1), 8),
            ('TEXTCOLOR',  (1, 1), (1, 1), colors.HexColor('#16A34A')),
            ('TEXTCOLOR',  (2, 1), (2, 1), colors.HexColor('#DC2626')),
        ]))
        elementos.append(resumen)

    elementos.append(Spacer(1, 0.3*cm))
    elementos.append(Paragraph('PISYS — Historial de movimientos', estilos['pie']))
    doc.build(elementos)
    return _enviar_pdf(buffer, 'historial.pdf')


# ---------------------------------------------------------------------------
# RUTAS — API BOT PREDICCIONES
# ---------------------------------------------------------------------------

@app.route('/api/predicciones')
@login_required
def api_predicciones():
    """
    Devuelve JSON con el análisis predictivo de todos los productos.
    El bot flotante consume este endpoint para generar sus respuestas.
    """
    from datetime import timedelta
    from collections import defaultdict
    import json

    hace_30 = datetime.utcnow() - timedelta(days=30)
    productos = Producto.query.all()

    resumen_inventario = {
        'total_productos':  len(productos),
        'valor_total':      round(sum(p.valor_total() for p in productos), 0),
        'criticos':         [],
        'bajos':            [],
        'normales':         [],
        'predicciones':     [],
        'sin_movimiento':   [],
    }

    for p in productos:
        info_base = {
            'id':        p.id,
            'nombre':    p.nombre,
            'categoria': p.categoria or 'General',
            'cantidad':  p.cantidad,
            'precio':    p.precio,
        }

        # Calcular promedio de salidas últimos 30 días
        salidas = [
            abs(m.cantidad_cambio)
            for m in p.movimientos
            if m.cantidad_cambio < 0 and m.fecha >= hace_30
        ]
        total_salidas = sum(salidas)
        promedio_diario = round(total_salidas / 30, 2)

        if promedio_diario > 0:
            dias_restantes = round(p.cantidad / promedio_diario, 1)
            fecha_agotamiento = (
                datetime.utcnow() + timedelta(days=dias_restantes)
            ).strftime('%d/%m/%Y')

            pred = {
                **info_base,
                'promedio_diario':    promedio_diario,
                'dias_restantes':     dias_restantes,
                'fecha_agotamiento':  fecha_agotamiento,
                'urgencia': (
                    'critica'      if dias_restantes <= 7  else
                    'advertencia'  if dias_restantes <= 14 else
                    'normal'
                ),
            }
            resumen_inventario['predicciones'].append(pred)
        else:
            resumen_inventario['sin_movimiento'].append(info_base)

        # Clasificar por estado actual
        estado = p.estado_stock()
        if estado == 'critico':
            resumen_inventario['criticos'].append(info_base)
        elif estado == 'bajo':
            resumen_inventario['bajos'].append(info_base)
        else:
            resumen_inventario['normales'].append(info_base)

    # Ordenar predicciones por urgencia
    resumen_inventario['predicciones'].sort(key=lambda x: x['dias_restantes'])

    from flask import jsonify
    return jsonify(resumen_inventario)


# ---------------------------------------------------------------------------
# RUTAS — DASHBOARD Y ALERTAS PREDICTIVAS (Semanas 10 y 11)
# ---------------------------------------------------------------------------

@app.route('/dashboard')
@requiere_rol('admin', 'auditor', 'operador')
@login_required
def dashboard():
    productos  = Producto.query.all()
    movimientos = Movimiento.query.order_by(Movimiento.fecha.desc()).all()

    # --- KPIs generales ---
    total_productos        = len(productos)
    valor_total_inventario = sum(p.valor_total() for p in productos)
    total_criticos         = len([p for p in productos if p.estado_stock() == 'critico'])
    total_movimientos      = len(movimientos)

    # --- Gráfico 1: stock por categoría ---
    # Agrupamos cantidad total por categoría para las barras
    from collections import defaultdict
    stock_por_cat = defaultdict(int)
    valor_por_cat = defaultdict(float)
    for p in productos:
        cat = p.categoria or 'General'
        stock_por_cat[cat] += p.cantidad
        valor_por_cat[cat] += p.valor_total()

    categorias_labels = list(stock_por_cat.keys())
    categorias_stock  = [stock_por_cat[c] for c in categorias_labels]
    categorias_valor  = [round(valor_por_cat[c], 0) for c in categorias_labels]

    # --- Gráfico 2: distribución de valor (pastel) ---
    pastel_labels = categorias_labels
    pastel_values = categorias_valor

    # --- Gráfico 3: movimientos por día (últimos 30 días) ---
    from datetime import timedelta
    from collections import Counter
    hace_30 = datetime.utcnow() - timedelta(days=30)
    movs_recientes = Movimiento.query.filter(Movimiento.fecha >= hace_30).all()

    entradas_por_dia = Counter()
    salidas_por_dia  = Counter()
    for m in movs_recientes:
        dia = m.fecha.strftime('%d/%m')
        if m.cantidad_cambio > 0:
            entradas_por_dia[dia] += m.cantidad_cambio
        else:
            salidas_por_dia[dia]  += abs(m.cantidad_cambio)

    # Generar eje X con todos los días del rango
    dias_labels = []
    for i in range(29, -1, -1):
        dia = (datetime.utcnow() - timedelta(days=i)).strftime('%d/%m')
        dias_labels.append(dia)
    entradas_vals = [entradas_por_dia.get(d, 0) for d in dias_labels]
    salidas_vals  = [salidas_por_dia.get(d, 0)  for d in dias_labels]

    # --- Semana 11: Alertas predictivas ---
    # Algoritmo simple: promedio de salidas diarias de los últimos 30 días.
    # Si el stock actual / promedio_diario < 7 días → alerta "se agota pronto".
    alertas_predictivas = []
    for p in productos:
        salidas_producto = [
            abs(m.cantidad_cambio)
            for m in p.movimientos
            if m.cantidad_cambio < 0
            and m.fecha >= hace_30
        ]
        if not salidas_producto:
            continue

        promedio_diario = sum(salidas_producto) / 30
        if promedio_diario <= 0:
            continue

        dias_restantes = p.cantidad / promedio_diario

        if dias_restantes <= 14:  # Menos de 2 semanas
            alertas_predictivas.append({
                'producto':       p,
                'dias_restantes': round(dias_restantes, 1),
                'promedio_diario': round(promedio_diario, 1),
                'urgencia':       'critica' if dias_restantes <= 7 else 'advertencia'
            })

    # Ordenar: primero los más urgentes
    alertas_predictivas.sort(key=lambda x: x['dias_restantes'])

    return render_template('dashboard.html',
        # KPIs
        total_productos        = total_productos,
        valor_total_inventario = valor_total_inventario,
        total_criticos         = total_criticos,
        total_movimientos      = total_movimientos,
        # Datos para gráficos (se pasan como listas Python → JSON en el template)
        categorias_labels = categorias_labels,
        categorias_stock  = categorias_stock,
        categorias_valor  = categorias_valor,
        pastel_labels     = pastel_labels,
        pastel_values     = pastel_values,
        dias_labels       = dias_labels,
        entradas_vals     = entradas_vals,
        salidas_vals      = salidas_vals,
        # Alertas predictivas
        alertas_predictivas = alertas_predictivas,
    )


# ---------------------------------------------------------------------------
# RUTAS — CARGA MASIVA DESDE EXCEL
# ---------------------------------------------------------------------------

@app.route('/carga-masiva')
@requiere_rol('admin', 'operador')
@login_required
def carga_masiva():
    return render_template('carga_masiva.html')


@app.route('/carga-masiva/plantilla')
@requiere_rol('admin', 'operador')
@login_required
def descargar_plantilla():
    """Genera y devuelve un archivo Excel listo para llenar."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Productos'

    # --- Estilos ---
    color_header   = '1E3A5F'
    color_ejemplo  = 'EFF6FF'
    borde_delgado  = Border(
        left   = Side(style='thin', color='CBD5E1'),
        right  = Side(style='thin', color='CBD5E1'),
        top    = Side(style='thin', color='CBD5E1'),
        bottom = Side(style='thin', color='CBD5E1'),
    )

    estilo_header = {
        'font':      Font(name='Calibri', bold=True, color='FFFFFF', size=11),
        'fill':      PatternFill(fill_type='solid', fgColor=color_header),
        'alignment': Alignment(horizontal='center', vertical='center', wrap_text=True),
        'border':    borde_delgado,
    }
    estilo_ejemplo = {
        'font':      Font(name='Calibri', color='374151', size=10),
        'fill':      PatternFill(fill_type='solid', fgColor=color_ejemplo),
        'alignment': Alignment(horizontal='left', vertical='center'),
        'border':    borde_delgado,
    }
    estilo_data = {
        'font':      Font(name='Calibri', size=10),
        'alignment': Alignment(horizontal='left', vertical='center'),
        'border':    borde_delgado,
    }

    # --- Encabezados ---
    columnas = [
        ('nombre',      'Nombre del producto *', 30),
        ('descripcion', 'Descripción',           40),
        ('categoria',   'Categoría *',           18),
        ('precio',      'Precio *',              14),
        ('cantidad',    'Cantidad en stock *',   20),
    ]

    for col_idx, (_, titulo, ancho) in enumerate(columnas, start=1):
        celda = ws.cell(row=1, column=col_idx, value=titulo)
        for attr, val in estilo_header.items():
            setattr(celda, attr, val)
        ws.column_dimensions[celda.column_letter].width = ancho

    ws.row_dimensions[1].height = 30

    # --- Fila de ejemplo ---
    ejemplo = ['Laptop Dell Inspiron 15', 'Intel i5, 8GB RAM, 256GB SSD',
               'Electrónica', 1850000, 12]
    for col_idx, valor in enumerate(ejemplo, start=1):
        celda = ws.cell(row=2, column=col_idx, value=valor)
        for attr, val in estilo_ejemplo.items():
            setattr(celda, attr, val)

    # --- 50 filas vacías para que el usuario llene ---
    for fila in range(3, 53):
        for col_idx in range(1, 6):
            celda = ws.cell(row=fila, column=col_idx, value='')
            for attr, val in estilo_data.items():
                setattr(celda, attr, val)
        ws.row_dimensions[fila].height = 18

    # --- Hoja de categorías válidas como referencia ---
    ws_cat = wb.create_sheet('Categorías válidas')
    ws_cat['A1'] = 'Categorías disponibles'
    ws_cat['A1'].font = Font(bold=True, color='FFFFFF', size=11)
    ws_cat['A1'].fill = PatternFill(fill_type='solid', fgColor=color_header)
    ws_cat['A1'].alignment = Alignment(horizontal='center')
    ws_cat.column_dimensions['A'].width = 25

    for i, cat in enumerate(CATEGORIAS, start=2):
        ws_cat.cell(row=i, column=1, value=cat)

    # --- Hoja de instrucciones ---
    ws_inst = wb.create_sheet('Instrucciones')
    instrucciones = [
        ('INSTRUCCIONES DE USO', True),
        ('', False),
        ('1. Llena la hoja "Productos" con tus productos.', False),
        ('2. La fila 2 (azul claro) es un ejemplo — puedes borrarla o reemplazarla.', False),
        ('3. Campos obligatorios: Nombre, Categoría, Precio y Cantidad.', False),
        ('4. La Categoría debe coincidir exactamente con las de la hoja "Categorías válidas".', False),
        ('5. El Precio debe ser un número (sin $, sin puntos de miles, sin comas).', False),
        ('   Ejemplo correcto: 1850000   Ejemplo incorrecto: $1.850.000', False),
        ('6. La Cantidad debe ser un número entero (sin decimales).', False),
        ('7. Filas completamente vacías se ignorarán automáticamente.', False),
        ('8. Si una fila tiene error, se reporta pero el resto sí se guarda.', False),
        ('', False),
        ('Máximo 500 productos por archivo.', True),
    ]
    ws_inst.column_dimensions['A'].width = 70
    for i, (texto, negrita) in enumerate(instrucciones, start=1):
        celda = ws_inst.cell(row=i, column=1, value=texto)
        celda.font = Font(bold=negrita, size=11 if negrita else 10)
        celda.alignment = Alignment(wrap_text=True)

    # --- Enviar como descarga ---
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return send_file(
        buffer,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='plantilla_pisys.xlsx'
    )


@app.route('/carga-masiva/procesar', methods=['POST'])
@requiere_rol('admin', 'operador')
@login_required
def procesar_carga_masiva():
    """Lee el Excel subido, valida cada fila y guarda los productos válidos."""

    archivo = request.files.get('archivo')

    if not archivo or archivo.filename == '':
        flash('❌ No seleccionaste ningún archivo.', 'danger')
        return redirect(url_for('carga_masiva'))

    if not archivo.filename.endswith(('.xlsx', '.xls')):
        flash('❌ El archivo debe ser Excel (.xlsx o .xls).', 'danger')
        return redirect(url_for('carga_masiva'))

    try:
        wb = openpyxl.load_workbook(archivo, data_only=True)
    except Exception:
        flash('❌ No se pudo leer el archivo. Asegúrate de que sea un Excel válido.', 'danger')
        return redirect(url_for('carga_masiva'))

    # Intentamos leer la hoja "Productos"; si no existe, tomamos la primera
    ws = wb['Productos'] if 'Productos' in wb.sheetnames else wb.active

    guardados  = []   # productos que se guardaron exitosamente
    errores    = []   # filas con problemas
    omitidas   = 0    # filas completamente vacías

    LIMITE = 500

    for num_fila, fila in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        # Ignorar filas completamente vacías
        if all(v is None or str(v).strip() == '' for v in fila):
            omitidas += 1
            continue

        if len(guardados) >= LIMITE:
            errores.append({'fila': num_fila, 'error': f'Límite de {LIMITE} productos alcanzado.'})
            continue

        # Leer cada columna con valor por defecto seguro
        nombre      = str(fila[0]).strip() if fila[0] is not None else ''
        descripcion = str(fila[1]).strip() if len(fila) > 1 and fila[1] is not None else ''
        categoria   = str(fila[2]).strip() if len(fila) > 2 and fila[2] is not None else 'General'
        precio_raw  = fila[3] if len(fila) > 3 else None
        cantidad_raw = fila[4] if len(fila) > 4 else None

        # --- Validaciones ---
        errores_fila = []

        if not nombre:
            errores_fila.append('Nombre vacío')

        if len(nombre) > 100:
            errores_fila.append('Nombre supera 100 caracteres')

        # Validar precio
        try:
            precio = float(str(precio_raw).replace(',', '.').replace('$', '').strip())
            if precio < 0:
                errores_fila.append('Precio negativo')
        except (ValueError, TypeError):
            precio = None
            errores_fila.append(f'Precio inválido: "{precio_raw}"')

        # Validar cantidad
        try:
            cantidad = int(float(str(cantidad_raw).strip()))
            if cantidad < 0:
                errores_fila.append('Cantidad negativa')
        except (ValueError, TypeError):
            cantidad = None
            errores_fila.append(f'Cantidad inválida: "{cantidad_raw}"')

        if errores_fila:
            errores.append({
                'fila':    num_fila,
                'nombre':  nombre or '(sin nombre)',
                'error':   ' · '.join(errores_fila)
            })
            continue

        # --- Guardar producto válido ---
        nuevo = Producto(
            nombre      = nombre,
            descripcion = descripcion,
            precio      = precio,
            cantidad    = cantidad,
            categoria   = categoria if categoria in CATEGORIAS else 'General',
        )
        db.session.add(nuevo)
        db.session.flush()  # obtener ID para el movimiento

        if nuevo.cantidad > 0:
            registrar_movimiento(nuevo, 0, tipo='creacion',
                                 nota='Carga masiva desde Excel')

        guardados.append(nombre)

    # Commit de todos los productos válidos de una sola vez
    if guardados:
        db.session.commit()

    # Guardar resultados en sesión para mostrarlos en la página
    resumen = {
        'guardados': guardados,
        'errores':   errores,
        'omitidas':  omitidas,
    }

    return render_template('carga_masiva.html', resumen=resumen)


# ---------------------------------------------------------------------------
# RUTAS — AUTENTICACIÓN
# ---------------------------------------------------------------------------

@app.route('/login', methods=['GET', 'POST'])
def login():
    if current_user.is_authenticated:
        return redirect(url_for('inicio'))

    if request.method == 'POST':
        email    = request.form.get('email', '').strip().lower()
        password = request.form.get('password', '')

        usuario = Usuario.query.filter_by(email=email).first()

        if not usuario or not usuario.check_password(password):
            flash('❌ Email o contraseña incorrectos.', 'danger')
            return render_template('login.html')

        if not usuario.activo:
            flash('❌ Tu cuenta está desactivada. Contacta al administrador.', 'danger')
            return render_template('login.html')

        login_user(usuario, remember=True)
        # Redirigir a la página que intentaba visitar, o al inicio
        siguiente = request.args.get('next') or url_for('inicio')
        return redirect(siguiente)

    return render_template('login.html')


@app.route('/logout')
@login_required
def logout():
    logout_user()
    flash('✅ Sesión cerrada correctamente.', 'success')
    return redirect(url_for('login'))


# ---------------------------------------------------------------------------
# RUTAS — GESTIÓN DE USUARIOS (solo admin)
# ---------------------------------------------------------------------------

@app.route('/usuarios')
@requiere_rol('admin', 'auditor')
@login_required
def lista_usuarios():
    usuarios = Usuario.query.order_by(Usuario.fecha_creacion.desc()).all()
    return render_template('usuarios.html', usuarios=usuarios)


@app.route('/usuarios/nuevo', methods=['GET', 'POST'])
@requiere_rol('admin')
@login_required
def nuevo_usuario():
    if request.method == 'POST':
        nombre   = request.form.get('nombre', '').strip()
        email    = request.form.get('email', '').strip().lower()
        password = request.form.get('password', '')
        rol      = request.form.get('rol', 'operador')

        errores = []
        if not nombre:   errores.append('El nombre es obligatorio.')
        if not email:    errores.append('El email es obligatorio.')
        if not password: errores.append('La contraseña es obligatoria.')
        if len(password) < 6: errores.append('La contraseña debe tener al menos 6 caracteres.')
        if rol not in ('admin', 'operador', 'auditor'):
            errores.append('Rol inválido.')
        if Usuario.query.filter_by(email=email).first():
            errores.append(f'El email {email} ya está registrado.')

        if errores:
            for e in errores: flash(f'❌ {e}', 'danger')
            return render_template('usuario_form.html', titulo='Nuevo usuario', usuario=None)

        u = Usuario(nombre=nombre, email=email, rol=rol)
        u.set_password(password)
        db.session.add(u)
        db.session.commit()
        flash(f'✅ Usuario "{nombre}" creado con rol {rol}.', 'success')
        return redirect(url_for('lista_usuarios'))

    return render_template('usuario_form.html', titulo='Nuevo usuario', usuario=None)


@app.route('/usuarios/editar/<int:id>', methods=['GET', 'POST'])
@requiere_rol('admin')
@login_required
def editar_usuario(id):
    usuario = Usuario.query.get_or_404(id)

    if request.method == 'POST':
        nombre   = request.form.get('nombre', '').strip()
        email    = request.form.get('email', '').strip().lower()
        rol      = request.form.get('rol', 'operador')
        password = request.form.get('password', '').strip()
        activo   = request.form.get('activo') == 'on'

        errores = []
        if not nombre: errores.append('El nombre es obligatorio.')
        if not email:  errores.append('El email es obligatorio.')
        if rol not in ('admin', 'operador', 'auditor'): errores.append('Rol inválido.')
        if password and len(password) < 6:
            errores.append('La contraseña debe tener al menos 6 caracteres.')

        # Verificar email único (excluyendo el propio usuario)
        existente = Usuario.query.filter_by(email=email).first()
        if existente and existente.id != id:
            errores.append(f'El email {email} ya está en uso.')

        # Evitar que el admin se desactive a sí mismo
        if usuario.id == current_user.id and not activo:
            errores.append('No puedes desactivar tu propia cuenta.')

        if errores:
            for e in errores: flash(f'❌ {e}', 'danger')
            return render_template('usuario_form.html',
                                   titulo=f'Editar: {usuario.nombre}', usuario=usuario)

        usuario.nombre = nombre
        usuario.email  = email
        usuario.rol    = rol
        usuario.activo = activo
        if password:
            usuario.set_password(password)

        db.session.commit()
        flash(f'✅ Usuario "{nombre}" actualizado.', 'success')
        return redirect(url_for('lista_usuarios'))

    return render_template('usuario_form.html',
                           titulo=f'Editar: {usuario.nombre}', usuario=usuario)


@app.route('/usuarios/eliminar/<int:id>', methods=['POST'])
@requiere_rol('admin')
@login_required
def eliminar_usuario(id):
    usuario = Usuario.query.get_or_404(id)

    if usuario.id == current_user.id:
        flash('❌ No puedes eliminar tu propia cuenta.', 'danger')
        return redirect(url_for('lista_usuarios'))

    nombre = usuario.nombre
    db.session.delete(usuario)
    db.session.commit()
    flash(f'🗑️ Usuario "{nombre}" eliminado.', 'warning')
    return redirect(url_for('lista_usuarios'))


# ---------------------------------------------------------------------------
# ARRANQUE — crea usuario admin por defecto si no existe ninguno
# ---------------------------------------------------------------------------

if __name__ == '__main__':
    with app.app_context():
        db.create_all()

        # Crear administrador por defecto si no hay ningún usuario
        if Usuario.query.count() == 0:
            admin = Usuario(nombre='Administrador', email='admin@pisys.com', rol='admin')
            admin.set_password('admin123')
            db.session.add(admin)
            db.session.commit()
            print('👤 Usuario admin creado: admin@pisys.com / admin123')
            print('   ⚠️  Cambia la contraseña después del primer login.')

        print('✅ Base de datos lista')
        print('🚀 http://localhost:5000')
    app.run(debug=True)
